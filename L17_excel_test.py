import openpyxl

# 엑셀 로드
wb = openpyxl.load_workbook('L17_test.xlsx')
print(type(wb))




# 시트 접근
for sheet_name in wb.sheetnames:  # sheet 목록 반환
    print(sheet_name)

act_sheet = wb.active   # 활성화된 시트 반환
print(act_sheet)


# 셀 접근 : cell() 함수
sheet = wb['Sheet1']
print(sheet.cell(row=1, column=2).value)

# 마지막셀 (row): sheet.max_row
# 마지막셀 (column): sheet.max_column
print(sheet.max_row)
print(sheet.max_column)

for i in range(1,8,2):  # 1,3, 5, 7
    print(sheet.cell(row=i, column=2))


# 여러셀 접근 : 범위지정
cell_range = sheet['A1':'C2']
colC = sheet['C']
col_range = sheet['C:D']
row10 = sheet[10]
row_range = sheet[5:10]

